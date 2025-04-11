#!/usr/bin/env python
import argparse
import sys
import os
import re
import traceback
import uuid
from pathlib import Path
import shutil
import copy
from io import BytesIO

# --- Dependency Imports ---
try:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.utils import get_column_letter
    from bs4 import BeautifulSoup, Tag, NavigableString
    from PIL import Image as PilImage
    try:
        from PIL import WmfImagePlugin, EmfImagePlugin # noqa F401
        print("DEBUG: Optional WMF/EMF support found in Pillow.", file=sys.stderr)
    except ImportError:
        print("DEBUG: Optional WMF/EMF support not found in Pillow.", file=sys.stderr)
        pass
    from dotenv import load_dotenv
    import mysql.connector
except ImportError as e:
    print(f"FATAL Error: Missing required Python package: {e.name}. "
          f"Please install dependencies from requirements.txt using "
          f"'pip install -r requirements.txt'", file=sys.stderr)
    sys.exit(1) # Exit immediately if dependencies are missing

# --- Configuration & Constants ---
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
DOTENV_PATH = PROJECT_ROOT / '.env'
STORAGE_APP_PATH = PROJECT_ROOT / 'storage' / 'app'
OUTPUTS_BASE_DIR_RELATIVE = Path('outputs')

# --- Logging ---
def log_debug(message): print(f"DEBUG: {message}", file=sys.stderr)
def log_info(message): print(f"INFO: {message}", file=sys.stderr)
def log_warning(message): print(f"WARNING: {message}", file=sys.stderr)
def log_error(message): print(f"ERROR: {message}", file=sys.stderr)

# --- Database Update Function ---
def update_job_status_db(job_id, status, output_path_abs=None, error_msg=None):
    # ... (Keep the robust version from previous responses) ...
    conn = None; cursor = None; relative_output_path = None
    try:
        if not load_dotenv(dotenv_path=DOTENV_PATH): log_warning(f"Could not load .env file from {DOTENV_PATH}.")
        db_host = os.getenv('DB_HOST', '127.0.0.1'); db_port = os.getenv('DB_PORT', '3306')
        db_name = os.getenv('DB_DATABASE'); db_user = os.getenv('DB_USERNAME'); db_pass = os.getenv('DB_PASSWORD')
        if not all([db_host, db_port, db_name, db_user]): print(f"CRITICAL: DB connection vars not set.", file=sys.stderr); return
        conn = mysql.connector.connect(host=db_host, port=int(db_port), database=db_name, user=db_user, password=db_pass, connect_timeout=10)
        cursor = conn.cursor()
        if output_path_abs:
            try:
                abs_storage_path = STORAGE_APP_PATH.resolve(); abs_output_path = Path(output_path_abs).resolve()
                if abs_storage_path in abs_output_path.parents: relative_output_path = abs_output_path.relative_to(abs_storage_path).as_posix()
                else: log_warning(f"Output path {abs_output_path} outside storage/app. Storing absolute."); relative_output_path = abs_output_path.as_posix()
            except ValueError as e: log_warning(f"Could not get relative path for {output_path_abs}. Storing absolute. Err: {e}"); relative_output_path = Path(output_path_abs).resolve().as_posix()
        if error_msg: error_msg = error_msg[:65530]
        sql = "UPDATE conversion_jobs SET status=%s, output_filepath=%s, error_message=%s, updated_at=NOW() WHERE id=%s AND status NOT IN ('completed', 'failed')"
        vals = (status, relative_output_path, error_msg, job_id)
        cursor.execute(sql, vals); conn.commit()
        log_info(f"DB Update: Job {job_id} status -> '{status}'. Rows: {cursor.rowcount}")
    except mysql.connector.Error as err: log_error(f"CRITICAL DB Update Error Job {job_id}: {err}")
    except Exception as e: log_error(f"CRITICAL General DB Update Error Job {job_id}: {e}\n{traceback.format_exc()}")
    finally:
        if cursor: cursor.close()
        if conn and conn.is_connected(): conn.close()

# --- Data Classes ---
class CellData:
    # ... (Keep CellData class same as previous version - Response #17) ...
    def __init__(self, cell: Cell | MergedCell | None, sheet: Worksheet, merged_cells_map: dict):
        self.cell = cell
        self.coordinate = cell.coordinate if cell else None
        self.row = cell.row if cell else None
        self.column = cell.column if cell else None
        self.is_merged = self.coordinate in merged_cells_map
        self.merge_range_info = merged_cells_map.get(self.coordinate)
        self.is_top_left_merged = self.is_merged and self.merge_range_info[0] == self.coordinate
        self.value = self._get_value(sheet)
        self.display_value = str(self.value) if self.value is not None else ""
        self.font = copy.copy(cell.font) if cell and hasattr(cell, 'font') and cell.font else None
        self.fill = copy.copy(cell.fill) if cell and hasattr(cell, 'fill') and cell.fill else None
        self.alignment = copy.copy(cell.alignment) if cell and hasattr(cell, 'alignment') and cell.alignment else None
        self.border = copy.copy(cell.border) if cell and hasattr(cell, 'border') and cell.border else None
        self.number_format = cell.number_format if cell and hasattr(cell, 'number_format') else 'General'
        self.rowspan = 1; self.colspan = 1
        if self.is_top_left_merged:
             range_obj = self.merge_range_info[1]
             self.rowspan = range_obj.max_row - range_obj.min_row + 1
             self.colspan = range_obj.max_col - range_obj.min_col + 1
    def _get_value(self, sheet: Worksheet):
        if self.cell is None: return None
        try:
            if self.is_merged and not self.is_top_left_merged: return sheet[self.merge_range_info[0]].value
            else: return self.cell.value
        except Exception: return None
    def is_empty(self): return self.value is None or (isinstance(self.value, str) and not self.value.strip())
    def get_formatted_value(self): return str(self.value) if self.value is not None else ""
    def is_bold(self): return self.font and self.font.bold
    def get_horizontal_alignment(self): return self.alignment.horizontal if self.alignment and self.alignment.horizontal else None
    def __str__(self): return self.display_value

class RowData:
    # ... (Keep RowData class same as previous version - Response #17) ...
    def __init__(self, row_idx: int, cells_data: list[CellData]):
        self.row_idx = row_idx; self.cells = cells_data
    def is_empty(self, ignore_merged_covered=True):
        for cell in self.cells:
             if ignore_merged_covered and cell.is_merged and not cell.is_top_left_merged: continue
             if not cell.is_empty(): return False
        return True
    def get_non_empty_cells(self, ignore_merged_covered=True):
        non_empty = []
        for cell in self.cells:
             if ignore_merged_covered and cell.is_merged and not cell.is_top_left_merged: continue
             if not cell.is_empty(): non_empty.append(cell)
        return non_empty
    def get_last_data_column_index(self):
         last_col = 0
         for cell in self.cells:
              if not cell.is_empty():
                   col_end = cell.column + (cell.colspan - 1 if cell.is_top_left_merged else 0)
                   last_col = max(last_col, col_end)
         return last_col
    def __len__(self): return len(self.cells)
    def __getitem__(self, index): return self.cells[index]

class ExcelSheetData:
    # ... (Keep ExcelSheetData class same as previous version - Response #17, including image extraction) ...
    def __init__(self, sheet: Worksheet):
        self.sheet = sheet; self.title = sheet.title; self.rows: list[RowData] = [];
        self.merged_cells_map: dict = self._create_merged_cells_map(); self.images: list[dict] = []; self._parse_sheet()
    def _create_merged_cells_map(self):
        merged_map = {};
        try:
            for merged_range in list(self.sheet.merged_cells.ranges):
                if not hasattr(merged_range, 'min_row'): continue
                min_col_ltr = get_column_letter(merged_range.min_col)
                top_left_coord = f"{min_col_ltr}{merged_range.min_row}"
                for row in range(merged_range.min_row, merged_range.max_row + 1):
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        coord = f"{get_column_letter(col)}{row}"; merged_map[coord] = (top_left_coord, merged_range)
        except Exception as e: log_error(f"Error creating merged cells map: {e}\n{traceback.format_exc()}")
        return merged_map
    def _parse_sheet(self):
        log_info(f"Parsing sheet '{self.title}'...")
        # *** CORRECTED: try: on a new line ***
        try:
            max_col_scan = 0;
            for row in self.sheet.iter_rows(): max_col_scan = max(max_col_scan, len(row))
            max_col = min(max_col_scan, self.sheet.max_column if self.sheet.max_column else max_col_scan)
            if max_col == 0: max_col = 50;
            if max_col > 150: log_warning(f"Limiting parsing to 150 columns (detected {max_col})."); max_col = 150
            min_col = self.sheet.min_column; log_info(f"Sheet dims: Max Col={max_col}, Max Row={self.sheet.max_row}")
            for row_idx, row_cells_tuple in enumerate(self.sheet.iter_rows(min_row=1, max_col=max_col, min_col=min_col), start=1):
                cell_data_list = [CellData(cell, self.sheet, self.merged_cells_map) for cell in row_cells_tuple]
                self.rows.append(RowData(row_idx, cell_data_list))
            log_info(f"Parsed {len(self.rows)} rows.")
        except Exception as e: log_error(f"Error during sheet parsing: {e}\n{traceback.format_exc()}"); raise
    def extract_images(self, output_image_dir_abs: Path, job_id: int):
        log_info("Extracting images..."); img_counter = 0;
        output_image_dir_abs.mkdir(parents=True, exist_ok=True);
        try: output_image_dir_rel = output_image_dir_abs.relative_to(STORAGE_APP_PATH / OUTPUTS_BASE_DIR_RELATIVE)
        except ValueError: output_image_dir_rel = Path(output_image_dir_abs.name); log_warning("Image dir not relative to outputs.")
        try:
            if not hasattr(self.sheet, '_images') or not self.sheet._images: log_info("No images found."); return
            for image in self.sheet._images:
                img_counter += 1; row_anchor, col_anchor = 1, 1
                try:
                    row_anchor = image.anchor._from.row + 1; col_anchor = image.anchor._from.col + 1
                    img_format = getattr(image, 'format', None) or getattr(image, 'contentType', '').split('/')[-1] or 'png'
                    suffix = f".{img_format.lower().split('+')[0]}"; img_filename = f"image_{job_id}_{img_counter}{suffix}"
                    img_abs_path = output_image_dir_abs / img_filename
                    img_rel_storage_path = (OUTPUTS_BASE_DIR_RELATIVE / output_image_dir_rel / img_filename).as_posix()
                    img_data = None
                    if hasattr(image, 'ref') and isinstance(image.ref, bytes): img_data = image.ref
                    elif hasattr(image, '_data') and callable(image._data): img_data = image._data()
                    if img_data:
                        pil_img = PilImage.open(BytesIO(img_data)); save_format = pil_img.format or img_format.upper() or 'PNG'
                        if save_format in ['WMF', 'EMF']: log_warning(f"Converting {save_format} to PNG (image {img_counter})."); save_format = 'PNG'; img_abs_path = img_abs_path.with_suffix('.png'); img_rel_storage_path = Path(img_rel_storage_path).with_suffix('.png').as_posix()
                        if pil_img.mode == 'P': pil_img = pil_img.convert('RGBA') if 'transparency' in pil_img.info else pil_img.convert('RGB')
                        elif pil_img.mode == 'RGBA' and save_format in ['JPEG', 'JPG']: pil_img = pil_img.convert('RGB')
                        pil_img.save(img_abs_path, format=save_format); log_info(f"Saved image: {img_abs_path} (Anchor: R{row_anchor}C{col_anchor})")
                        self.images.append({"row": row_anchor, "col": col_anchor, "path": img_rel_storage_path})
                    else: log_warning(f"No data bytes for image {img_counter} @ R{row_anchor}C{col_anchor}.")
                except Exception as img_ex: log_error(f"Failed processing image {img_counter} @ R{row_anchor}C{col_anchor}: {img_ex}")
        except Exception as e: log_error(f"General image extraction error: {e}\n{traceback.format_exc()}")
        log_info(f"Image extraction finished. Saved {len(self.images)} images.")
    def get_rows_in_range(self, start_row, end_row):
        if start_row < 1 or end_row > len(self.rows): return []
        return self.rows[start_row-1 : end_row]

# --- HTML Generation Logic (Refactored Stage 1) ---
BOOTSTRAP_HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-rbsA2VBKQhggwzxH7pPCaAqO46MgnOM80zW1RWuH61DGLwZJEdK2Kadq2F9CUG65" crossorigin="anonymous">
    <style>
        body {{ font-family: sans-serif; padding-top: 1rem; font-size: 0.9rem; background-color: #f8f9fa; }}
        .excel-container {{ max-width: 1200px; margin: auto; background-color: #fff; padding: 1.5rem; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
        .cable-title {{ font-size: 1.8rem; font-weight: bold; margin: 1rem 0 1.5rem 0; color: #004085; text-align: center; }}
        .section-header {{ font-size: 1.3rem; font-weight: bold; margin-top: 2rem; margin-bottom: 1rem; padding-bottom: 0.5rem; border-bottom: 2px solid #dee2e6; color: #495057; text-align: center; }}
        .section-description {{ margin-bottom: 1rem; text-align: justify; line-height: 1.6; color: #333; }}
        .table-responsive {{ margin-bottom: 1.5rem; border: 1px solid #dee2e6; border-radius: 5px; overflow-x: auto; }}
        .table {{ margin-bottom: 0; }}
        .table thead th {{ background-color: #e9ecef; font-weight: bold; text-align: center; vertical-align: middle; border-color: #dee2e6; border-bottom-width: 2px; padding: 0.5rem 0.6rem;}}
        .table tbody th {{ background-color: #f8f9fa; font-weight: 600; text-align: left; border-color: #dee2e6; padding: 0.4rem 0.5rem; }} /* Row header style */
        .table td, .table th {{ vertical-align: middle; text-align: left; border-color: #dee2e6; padding: 0.4rem 0.5rem; }}
        .table td.text-center, .table th.text-center {{ text-align: center; }}
        .table-bordered > :not(caption) > * > * {{ border-width: 1px; }}
        /* Remove table-striped background by default for consistency */
        /* .table-striped > tbody > tr:nth-of-type(odd) > * {{ background-color: rgba(0, 0, 0, 0.03); }} */
        img.excel-image {{ max-width: 90%; height: auto; display: block; margin: 1.5rem auto; border: 1px solid #eee; border-radius: 4px; padding: 5px; }}
        .footer-text {{ margin-top: 2rem; padding: 1rem; background-color: #f8f9fa; border-radius: 5px; font-size: 0.85rem; color: #6c757d; border-left: 3px solid #004085; }}
    </style>
</head>
<body>
    <div class="container excel-container">
        {content}
    </div>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/js/bootstrap.bundle.min.js" integrity="sha384-kenU1KFdBIe4zVF0s0G1M5b4hcpxyD9F7jL+jjXkk+Q2h455rYXK/7HAuoJl+0I4" crossorigin="anonymous"></script>
</body>
</html>"""

class HtmlConverter:
    """Converts ExcelSheetData to structured HTML (Stage 1 Refactor)."""

    SECTION_HEADERS = [
        "APPLICATION", "STANDARDS", "CHARACTERISTICS", "CONSTRUCTION",
        "DIMENSIONS", "ELECTRICAL CHARACTERISTICS", "CONDUCTORS",
        "DE-RATING", "CURRENT CARRYING", "VOLTAGE DROP", "FACTORS", "FEATURES", "TECHNICAL DATA"
    ]
    SECTION_HEADERS_UPPER = [h.upper() for h in SECTION_HEADERS]
    DESCRIPTIVE_KEYWORDS_LOWER = ['for use in', 'according to', 'conductor operates', 'note', 'where a conductor', 'fixed wiring', 'power networks']
    FOOTER_KEYWORDS_LOWER = ["information contained", "guidance only", "subject to change", "liability", "tolerances", "manufacturing tolerances"]

    def __init__(self, sheet_data: ExcelSheetData, job_id: int):
        self.sheet_data = sheet_data
        self.job_id = job_id
        self.soup = BeautifulSoup("", "html.parser")
        self.content_elements: list[Tag | NavigableString] = []
        self.images_used: set[str] = set()

    # --- Row Classification (Improved Logic) ---
    def _classify_row(self, row_data: RowData, prev_classification: str | None) -> str:
        if row_data.is_empty(): return 'empty'
        non_empty_cells = row_data.get_non_empty_cells()
        if not non_empty_cells: return 'empty'

        first_cell = non_empty_cells[0]
        first_text = str(first_cell.value).strip()
        first_text_lower = first_text.lower()
        first_text_upper = first_text.upper()
        first_cell_col = first_cell.column
        min_sheet_col = self.sheet_data.sheet.min_column

        # --- Check Priority Order ---

        # 1. Title Check (Only very beginning)
        if row_data.row_idx <= 3 and prev_classification is None and self._is_like_title(row_data, first_text):
            return 'title'

        # 2. Footer Check (Only near end)
        if row_data.row_idx >= len(self.sheet_data.rows) - 5 and self._is_like_footer(row_data, first_text_lower):
             return 'footer'

        # 3. Section Header Check
        if first_cell_col == min_sheet_col and self._is_like_section_header(row_data, first_text_upper, first_text_lower):
             return 'section_header'

        # 4. Table Row Check (More likely if previous was table or header)
        has_borders = first_cell.border and (first_cell.border.left or first_cell.border.right or first_cell.border.top or first_cell.border.bottom)
        is_continuation = prev_classification in ['table_row', 'section_header']
        if len(non_empty_cells) > 1 or \
           (len(non_empty_cells) == 1 and first_cell.colspan > 1) or \
           has_borders or \
           is_continuation:
             # Check if it looks more like a description paragraph despite criteria
             if len(non_empty_cells) == 1 and first_cell.colspan > 1 and len(first_text.split()) > 10 and not has_borders and not is_continuation:
                  # Single, merged cell with long text, not following header/table -> likely paragraph
                  return 'paragraph_line'
             return 'table_row'

        # 5. Default: Paragraph Line
        # If only one cell, doesn't span, no borders, not header/footer/title -> paragraph
        return 'paragraph_line'

    def _is_like_title(self, row_data: RowData, first_text: str) -> bool:
        cell = row_data.cells[0]
        keywords = ['cable', 'conductor', 'wire']
        voltage_pattern = r'\d+(\.\d+)?\s*/\s*\d+(\.\d+)?\s*k?v'
        is_long_and_spans = len(first_text) > 15 and cell.colspan > max(1, len(row_data.cells) // 2)
        has_keywords = any(k in first_text.lower() for k in keywords) and re.search(voltage_pattern, first_text, re.IGNORECASE)
        is_basec = 'basec' in first_text.lower()
        # Title usually has large font, check if possible?
        # is_large_font = cell.font and cell.font.sz and cell.font.sz > 16

        return is_long_and_spans or has_keywords or is_basec # Add is_large_font if reliable

    def _is_like_section_header(self, row_data: RowData, first_text_upper: str, first_text_lower: str) -> bool:
        # Condition: Starts in first column, contains known header text,
        # AND (is the only significant content OR spans widely)
        # AND is not long descriptive text
        if not any(header in first_text_upper for header in self.SECTION_HEADERS_UPPER):
            return False
        # Avoid classifying descriptions as headers
        if len(first_text_upper.split()) > 8 or '\n' in first_text_upper or \
           any(kw in first_text_lower for kw in self.DESCRIPTIVE_KEYWORDS_LOWER):
            return False

        first_cell = row_data.cells[0]
        spans_wide = first_cell.colspan >= max(2, len(row_data.cells) // 2)
        only_significant = True
        for i, cell in enumerate(row_data.cells):
            if i == 0: continue
            # Check if cell is *visibly* covered by the first cell's merge span
            is_covered = False
            if first_cell.is_merged:
                 range_obj = first_cell.merge_range_info[1]
                 if cell.row == first_cell.row and cell.column >= range_obj.min_col and cell.column <= range_obj.max_col:
                      is_covered = True

            if not is_covered and not cell.is_empty() and len(str(cell.value).strip()) > 5: # Allow short codes
                 only_significant = False
                 break
        return spans_wide or only_significant

    def _is_like_footer(self, row_data: RowData, first_text_lower: str) -> bool:
        # Condition: Contains footer keywords AND is likely the only content in the row
        if len(row_data.get_non_empty_cells(ignore_merged_covered=False)) == 1 and \
           any(k in first_text_lower for k in self.FOOTER_KEYWORDS_LOWER):
            return True
        return False

    # --- Element Creation Helpers ---
    # _add_image_tag, _add_title_tag, _add_section_header_tag, _add_footer_tag, _append_text_with_br
    # Keep these largely the same as previous version (Response #17)
    def _add_image_tag(self, row_idx):
        for img_info in self.sheet_data.images:
            if img_info['row'] == row_idx and img_info['path'] not in self.images_used:
                public_path = f"/storage/{img_info['path']}"
                img_tag = self.soup.new_tag('img', attrs={'src': public_path, 'class': 'excel-image img-fluid rounded my-3', 'alt': f'Image from Excel near row {row_idx}', 'loading': 'lazy'})
                self.images_used.add(img_info['path']); log_debug(f"Prepared image tag for row {row_idx}: {public_path}"); return img_tag
        return None
    def _add_title_tag(self, row_data: RowData):
        title_cell = next((cell for cell in row_data.cells if not cell.is_empty()), None)
        if title_cell:
            title_text = str(title_cell.value).strip(); title_tag = self.soup.new_tag('h1', attrs={'class': 'cable-title'})
            self._append_text_with_br(title_tag, title_text); self.content_elements.append(title_tag); log_info(f"Added Title: {title_text[:50]}...")
    def _add_section_header_tag(self, row_data: RowData):
        header_cell = next((cell for cell in row_data.cells if not cell.is_empty()), None)
        if header_cell:
            header_text = str(header_cell.value).strip(); header_tag = self.soup.new_tag('h2', attrs={'class': 'section-header'})
            header_tag.string = header_text; self.content_elements.append(header_tag); log_info(f"Added Section Header: {header_text}")
    def _add_footer_tag(self, row_data: RowData):
        footer_cell = next((cell for cell in row_data.cells if not cell.is_empty()), None)
        if footer_cell:
            footer_text = str(footer_cell.value).strip(); footer_div = self.soup.new_tag('div', attrs={'class':'footer-text'})
            self._append_text_with_br(footer_div, footer_text); self.content_elements.append(footer_div); log_info("Added Footer text.")
    def _append_text_with_br(self, parent_tag: Tag, text: str):
         if text is None: return
         lines = str(text).split('\n');
         for i, line in enumerate(lines):
              line = line.strip();
              if line or i < len(lines) - 1:
                   parent_tag.append(NavigableString(line))
                   if i < len(lines) - 1: parent_tag.append(self.soup.new_tag('br'))

    # --- Table Creation (Refined - Stage 1) ---
    def _create_table_element(self, table_rows_data: list[RowData]) -> Tag | None:
        """Creates a structured HTML table (Stage 1 Refactor: Thead/Th improved, No striped)."""
        if not table_rows_data: return None

        # REMOVED 'table-striped' class
        table_tag = self.soup.new_tag('table', attrs={'class': 'table table-bordered table-sm'})
        tbody_tag = self.soup.new_tag('tbody')
        processed_coords_in_rowspan = set()

        max_col_index = 0
        for row_data in table_rows_data: max_col_index = max(max_col_index, row_data.get_last_data_column_index())
        if max_col_index == 0: return None
        log_debug(f"Table block (start row {table_rows_data[0].row_idx}): Max Col = {max_col_index}")

        # --- Identify Header Rows (Improved Heuristic) ---
        header_rows_data = []
        data_rows_data = []
        potential_header_end_idx = 0
        for idx, row_data in enumerate(table_rows_data):
            non_empty = row_data.get_non_empty_cells(ignore_merged_covered=False) # Consider all non-empty
            if not non_empty: continue

            # Header if: All visible cells are bold OR complex merges exist in early rows OR background?
            all_bold = all(cell.is_bold() for cell in non_empty)
            has_complex_merge = any(cell.rowspan > 1 or cell.colspan > 1 for cell in non_empty)

            # More flexible header detection, allowing non-bold if structure is complex early on
            if (idx < 3 and (all_bold or has_complex_merge)) or \
               (idx >= 3 and all_bold): # Bold rows later might still be headers
                 header_rows_data.append(row_data)
                 potential_header_end_idx = idx + 1
                 log_debug(f"Row {row_data.row_idx} identified as potential header row.")
            else:
                 break # Stop header collection
        data_rows_data = table_rows_data[potential_header_end_idx:]

        # --- Build Thead ---
        if header_rows_data:
            thead_tag = self.soup.new_tag('thead')
            table_tag.append(thead_tag)
            thead_processed_coords = set() # Track rowspan within thead itself
            for row_data in header_rows_data:
                tr = self.soup.new_tag('tr')
                for cell_data in row_data.cells:
                    if cell_data.column > max_col_index: continue
                    current_coord = cell_data.coordinate
                    if current_coord in thead_processed_coords: continue
                    if cell_data.is_merged and not cell_data.is_top_left_merged: continue

                    th = self.soup.new_tag('th') # Always TH in thead
                    if cell_data.colspan > 1: th['colspan'] = str(cell_data.colspan)
                    if cell_data.rowspan > 1: th['rowspan'] = str(cell_data.rowspan)
                    self._append_text_with_br(th, cell_data.get_formatted_value())

                    css_classes = [] # Add styling based on alignment etc.
                    h_align = cell_data.get_horizontal_alignment();
                    if h_align == 'center': css_classes.append('text-center')
                    elif h_align == 'right': css_classes.append('text-end')
                    if css_classes: th['class'] = " ".join(css_classes)
                    tr.append(th)

                    if cell_data.rowspan > 1: # Mark for future rows within thead
                        for i in range(1, cell_data.rowspan):
                             for j in range(cell_data.colspan):
                                 thead_processed_coords.add(f"{get_column_letter(cell_data.column + j)}{row_data.row_idx + i}")
                if tr.find('th'): thead_tag.append(tr)
            # Add coords covered by thead rowspans to the main set
            processed_coords_in_rowspan.update(thead_processed_coords)

        # --- Build Tbody ---
        table_tag.append(tbody_tag)
        for row_data in data_rows_data:
            tr = self.soup.new_tag('tr')
            is_first_visible_cell_in_row = True
            for cell_data in row_data.cells:
                if cell_data.column > max_col_index: continue
                current_coord = cell_data.coordinate
                if current_coord in processed_coords_in_rowspan: continue
                if cell_data.is_merged and not cell_data.is_top_left_merged: continue

                # Use TH for first *visible* cell in tbody row
                is_row_header = is_first_visible_cell_in_row
                cell_tag_name = 'th' if is_row_header else 'td'
                cell_tag = self.soup.new_tag(cell_tag_name)
                if is_row_header: cell_tag['scope'] = 'row'

                if cell_data.colspan > 1: cell_tag['colspan'] = str(cell_data.colspan)
                if cell_data.rowspan > 1: cell_tag['rowspan'] = str(cell_data.rowspan)

                self._append_text_with_br(cell_tag, cell_data.get_formatted_value())

                css_classes = []
                h_align = cell_data.get_horizontal_alignment()
                if h_align == 'center': css_classes.append('text-center')
                elif h_align == 'right': css_classes.append('text-end')
                if css_classes: cell_tag['class'] = " ".join(css_classes)

                tr.append(cell_tag)
                if cell_data.rowspan > 1:
                    for i in range(1, cell_data.rowspan):
                         for j in range(cell_data.colspan):
                             processed_coords_in_rowspan.add(f"{get_column_letter(cell_data.column + j)}{row_data.row_idx + i}")
                is_first_visible_cell_in_row = False

            if tr.find(['td', 'th']): tbody_tag.append(tr) # Add only if row has cells

        if not table_tag.find('tr'): return None # Return None if table is empty
        return table_tag

    # --- Flushing and Main Conversion Logic (Improved Classification) ---
    def _flush_table(self, current_table_rows: list[RowData]):
        if not current_table_rows: return
        log_debug(f"Flushing table buffer ({len(current_table_rows)} rows, start row {current_table_rows[0].row_idx}).")
        table_element = self._create_table_element(current_table_rows)
        if table_element:
            responsive_div = self.soup.new_tag('div', attrs={'class': 'table-responsive'})
            responsive_div.append(table_element)
            self.content_elements.append(responsive_div)
            log_debug(f"Added flushed table to content.")
        else: log_debug(f"Flushed table was empty or invalid.")

    def _flush_paragraph(self, current_paragraph_lines: list[str]):
        if not current_paragraph_lines: return
        para_text = "\n".join(current_paragraph_lines).strip()
        if para_text:
            para_tag = self.soup.new_tag('p', attrs={'class':'section-description'})
            self._append_text_with_br(para_tag, para_text)
            self.content_elements.append(para_tag)
            log_debug(f"Added flushed paragraph: {para_text[:60]}...")

    def convert(self) -> str:
        """Performs the conversion using improved row classification."""
        log_info("Starting structured HTML conversion process (Stage 1 Refactor)...")
        current_table_rows = []
        current_paragraph_lines = []
        last_classification = None
        title_image_tag = None # Store potential title image

        # Find potential title image early
        for row_data in self.sheet_data.rows[:5]: # Check first few rows
             img_tag = self._add_image_tag(row_data.row_idx)
             if img_tag:
                 title_image_tag = img_tag # Assume first image is title image
                 break

        for row_data in self.sheet_data.rows:
            classification = self._classify_row(row_data, last_classification)
            log_debug(f"Row {row_data.row_idx}: Classified as '{classification}'")

            # Get image for this row (if not already used as title image)
            img_tag = None
            if not title_image_tag or row_data.row_idx != title_image_tag.get('data-row-idx', -1): # Hacky way to check if it's the title image row
                 img_tag = self._add_image_tag(row_data.row_idx)


            # --- State Machine ---
            # Finish previous block if context changes
            if classification != 'table_row' and last_classification == 'table_row':
                self._flush_table(current_table_rows); current_table_rows = []
            if classification == 'table_row' and last_classification == 'paragraph_line':
                 self._flush_paragraph(current_paragraph_lines); current_paragraph_lines = []
            if classification == 'paragraph_line' and last_classification == 'table_row':
                 self._flush_table(current_table_rows); current_table_rows = []
            # Add more state transitions as needed

            # --- Handle Image Placement ---
            # Add image *before* its associated content (except title)
            if img_tag and classification != 'title':
                 # If starting a new non-paragraph block, flush previous paragraph
                 if classification != 'paragraph_line' and last_classification == 'paragraph_line':
                      self._flush_paragraph(current_paragraph_lines); current_paragraph_lines = []
                 self.content_elements.append(img_tag)


            # --- Handle Current Row Classification ---
            if classification == 'title':
                self._flush_paragraph(current_paragraph_lines); self._flush_table(current_table_rows)
                current_paragraph_lines, current_table_rows = [], []
                self._add_title_tag(row_data)
                # Add the pre-found title image *after* the title H1
                if title_image_tag:
                    self.content_elements.append(title_image_tag)
            elif classification == 'section_header':
                self._flush_paragraph(current_paragraph_lines); self._flush_table(current_table_rows)
                current_paragraph_lines, current_table_rows = [], []
                self._add_section_header_tag(row_data)
                # Image was already added before if it belongs here
            elif classification == 'footer':
                self._flush_paragraph(current_paragraph_lines); self._flush_table(current_table_rows)
                current_paragraph_lines, current_table_rows = [], []
                self._add_footer_tag(row_data)
            elif classification == 'table_row':
                current_table_rows.append(row_data)
            elif classification == 'paragraph_line':
                line_text = " ".join(str(c.value) for c in row_data.get_non_empty_cells()).strip()
                if line_text: current_paragraph_lines.append(line_text)
            elif classification == 'empty':
                 if last_classification == 'table_row': self._flush_table(current_table_rows); current_table_rows = []
                 # Treat empty row as paragraph break too
                 if last_classification == 'paragraph_line': self._flush_paragraph(current_paragraph_lines); current_paragraph_lines = []
                 pass # Ignore empty rows otherwise

            # Update last classification (ignore empty)
            if classification != 'empty':
                last_classification = classification

        # Flush remaining buffers
        self._flush_table(current_table_rows)
        self._flush_paragraph(current_paragraph_lines)

        # Assemble final HTML
        content_html = "".join(str(elem) for elem in self.content_elements)
        final_html = BOOTSTRAP_HTML_TEMPLATE.format(title=self.sheet_data.title, content=content_html)
        log_info("HTML conversion finished.")
        return final_html


# --- SQL Generation Logic ---
class SqlConverter:
    # ... (Keep improved SqlConverter logic from previous version - Response #17) ...
    def __init__(self, sheet_data: ExcelSheetData, job_id):
         self.sheet_data = sheet_data; self.job_id = job_id
    def convert(self, target_section="DIMENSIONS"):
        sql_statements = []; log_info(f"Processing sheet '{self.sheet_data.title}' for SQL (target: {target_section})...")
        target_table_rows = []; in_target_section = False; section_header_upper = target_section.upper()
        temp_html_converter = HtmlConverter(self.sheet_data, self.job_id) # For classifier
        for row_data in self.sheet_data.rows:
            classification = temp_html_converter._classify_row(row_data, 'start')
            if classification == 'section_header':
                header_text = str(row_data.cells[0].value).strip().upper()
                if section_header_upper in header_text: in_target_section = True; log_info(f"Found target SQL section '{target_section}'.")
                elif in_target_section: log_info(f"End of target SQL section detected."); break
            elif in_target_section and classification == 'table_row': target_table_rows.append(row_data)
            elif in_target_section and classification not in ['table_row', 'empty']: log_info(f"End of SQL table data due to '{classification}'."); break
        if not target_table_rows: log_warning(f"No table rows for SQL section '{target_section}'."); return f"-- No table data for {target_section}.\n"
        header_row_data = None; data_start_index = 0
        for i, row_data in enumerate(target_table_rows):
             if len(row_data.get_non_empty_cells(ignore_merged_covered=False)) > 0: header_row_data = row_data; data_start_index = i + 1; break
        if not header_row_data: log_warning("No header row for SQL."); return "-- No header row.\n"
        max_col_sql = 0;
        for r_data in target_table_rows: max_col_sql = max(max_col_sql, r_data.get_last_data_column_index())
        headers_raw = [];
        for cell in header_row_data.cells:
             if cell.column > max_col_sql: break
             if not cell.is_merged or cell.is_top_left_merged: headers_raw.append(cell.value)
        headers = [];
        for h in headers_raw:
             if h is None: continue; s = re.sub(r'\s+', '_', str(h)).strip(); s = re.sub(r'[^a-zA-Z0-9_]', '', s).lower();
             if s: headers.append(s)
        if not headers: log_warning("No valid SQL headers."); return "-- No valid headers.\n"; log_info(f"SQL Headers: {headers}")
        table_name = f"excel_import_job_{self.job_id}_{target_section.lower()}"[:64]
        create_sql = f"CREATE TABLE IF NOT EXISTS `{table_name}` (\n  `import_id` INT AUTO_INCREMENT PRIMARY KEY,\n";
        for h in headers: create_sql += f"  `{h}` TEXT NULL,\n"
        create_sql += "  `source_row_index` INT NULL,\n  `imported_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP\n);\n"
        sql_statements.append(create_sql); sql_statements.append("\n-- Data Inserts --\n")
        insert_base = f"INSERT INTO `{table_name}` (`{'`, `'.join(headers)}`, `source_row_index`) VALUES "; values = []
        for row_data in target_table_rows[data_start_index:]:
            if row_data.is_empty(): continue; row_vals = []; col_count = 0; cell_idx = 0
            while col_count < len(headers) and cell_idx < len(row_data.cells):
                 cell = row_data.cells[cell_idx];
                 if not cell.is_merged or cell.is_top_left_merged:
                     v = cell.value; row_vals.append("NULL" if v is None else f"'{str(v).replace(chr(92), chr(92)*2).replace(chr(39), chr(39)*2)}'") # Basic escaping '\\' and '\''
                     col_count += cell.colspan
                 cell_idx += 1
            while len(row_vals) < len(headers): row_vals.append("NULL")
            row_vals.append(str(row_data.row_idx)); values.append(f"({', '.join(row_vals)})")
        chunk_size = 100
        for i in range(0, len(values), chunk_size):
             chunk = values[i:i + chunk_size];
             if chunk: sql_statements.append(insert_base + ",\n".join(chunk) + ";\n")
        return "".join(sql_statements)


# --- Main Execution ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Process Excel File to HTML or SQL.')
    parser.add_argument('input_path', help='Absolute Path to the input Excel file')
    parser.add_argument('output_path', help='Absolute Path to save the output file')
    parser.add_argument('job_id', help='Job ID for status updates', type=int)
    parser.add_argument('format', help='Output format (html/sql)', default='html', choices=['html', 'sql'])
    parser.add_argument('--sql-target-section', default='DIMENSIONS', help='Section header to target for SQL extraction')
    args = parser.parse_args()

    log_info(f"----- Python Script Start -----"); log_info(f"Job ID: {args.job_id}, Format: {args.format}")
    log_info(f"Input: {args.input_path}"); log_info(f"Output: {args.output_path}"); log_info(f"Env: {DOTENV_PATH}")

    output_dir_abs = Path(args.output_path).parent; output_dir_abs.mkdir(parents=True, exist_ok=True)

    try:
        input_file = Path(args.input_path)
        if not input_file.is_file(): raise FileNotFoundError(f"Input file not found: {args.input_path}")
        if input_file.stat().st_size == 0: raise ValueError(f"Input file is empty: {args.input_path}")

        log_info("Loading workbook...");
        try: workbook: Workbook = openpyxl.load_workbook(args.input_path, data_only=True, read_only=False)
        except Exception as load_err: raise ValueError(f"Failed load Excel (check format/corruption). Err: {load_err}")

        if not workbook.active: raise ValueError("Workbook has no active sheet.")
        log_info(f"Processing sheet: {workbook.active.title}"); sheet_data = ExcelSheetData(workbook.active)

        output_image_dir_abs = output_dir_abs / f"job_{args.job_id}_images"; sheet_data.extract_images(output_image_dir_abs, args.job_id)

        output_content = None
        if args.format == 'html':
            log_info("Converting to HTML..."); converter = HtmlConverter(sheet_data, args.job_id); output_content = converter.convert()
        elif args.format == 'sql':
            log_info(f"Converting to SQL (target: {args.sql_target_section})..."); converter = SqlConverter(sheet_data, args.job_id); output_content = converter.convert(target_section=args.sql_target_section)

        if not output_content: raise ValueError(f"Empty content after conversion (format: {args.format}).")

        log_info(f"Writing output: {args.output_path}")
        try:
            with open(args.output_path, 'w', encoding='utf-8') as f: f.write(output_content)
            log_info("Output written successfully.")
        except Exception as write_err: raise IOError(f"Failed to write output file {args.output_path}. Err: {write_err}")

        if args.format == 'sql' and output_image_dir_abs.exists():
             try: shutil.rmtree(output_image_dir_abs); log_info(f"Removed image dir: {output_image_dir_abs}")
             except Exception as rm_err: log_warning(f"Could not remove image dir {output_image_dir_abs}: {rm_err}")

        update_job_status_db(args.job_id, 'completed', args.output_path); log_info(f"----- Python Script End (Success) -----"); sys.exit(0)

    except (FileNotFoundError, ValueError, IOError) as known_error:
        error_message = f"{type(known_error).__name__}: {str(known_error)}"; log_error(error_message); update_job_status_db(args.job_id, 'failed', error_msg=error_message); sys.exit(1)
    except Exception as e:
        error_message = f"Unexpected Python Error: {str(e)}"; log_error(error_message + f"\n{traceback.format_exc()}"); update_job_status_db(args.job_id, 'failed', error_msg=error_message[:500]); sys.exit(1)