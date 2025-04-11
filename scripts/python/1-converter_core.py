#!/usr/bin/env python
import argparse
import sys
import os
import re
import traceback
import uuid
from pathlib import Path
import shutil

# --- Dependency Imports ---
try:
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from bs4 import BeautifulSoup, Tag, NavigableString
    from PIL import Image as PilImage
    from dotenv import load_dotenv
    import mysql.connector
except ImportError as e:
    print(f"Error: Missing required Python package: {e.name}. Please run 'pip install -r requirements.txt'", file=sys.stderr)
    sys.exit(1)


# --- Configuration & Constants ---
# Load environment variables from Laravel's .env file
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent # Assumes scripts/python is one level down from project root
DOTENV_PATH = PROJECT_ROOT / '.env'

STORAGE_APP_PATH = PROJECT_ROOT / 'storage' / 'app'
OUTPUTS_BASE_DIR_RELATIVE = 'outputs' # Relative to storage/app

# --- Logging Setup (Optional but Recommended) ---
# import logging
# logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', stream=sys.stderr)
# logger = logging.getLogger(__name__)
# Use print statements to stderr for simplicity in this context
def log_debug(message): print(f"DEBUG: {message}", file=sys.stderr)
def log_info(message): print(f"INFO: {message}", file=sys.stderr)
def log_warning(message): print(f"WARNING: {message}", file=sys.stderr)
def log_error(message): print(f"ERROR: {message}", file=sys.stderr)


# --- Database Update Function ---
def update_job_status_db(job_id, status, output_path_abs=None, error_msg=None):
    """Updates the job status in the MySQL database."""
    conn = None
    cursor = None
    relative_output_path = None
    try:
        if not load_dotenv(dotenv_path=DOTENV_PATH):
            log_warning(f"Could not load .env file from {DOTENV_PATH}. Attempting to use environment variables directly.")

        db_host = os.getenv('DB_HOST', '127.0.0.1')
        db_port = os.getenv('DB_PORT', '3306')
        db_name = os.getenv('DB_DATABASE')
        db_user = os.getenv('DB_USERNAME')
        db_pass = os.getenv('DB_PASSWORD')

        if not all([db_host, db_port, db_name, db_user]):
            raise ValueError("Database environment variables not fully set (DB_HOST, DB_PORT, DB_DATABASE, DB_USERNAME).")

        log_debug(f"Connecting to DB: Host={db_host}, Port={db_port}, DB={db_name}, User={db_user}")
        conn = mysql.connector.connect(
            host=db_host,
            port=int(db_port), # Ensure port is integer
            database=db_name,
            user=db_user,
            password=db_pass,
            connect_timeout=10 # Add a connection timeout
        )
        cursor = conn.cursor()

        # Convert absolute output path to relative path (relative to storage/app)
        if output_path_abs:
            try:
                # Ensure both paths are absolute before calculating relative path
                abs_storage_path = STORAGE_APP_PATH.resolve()
                abs_output_path = Path(output_path_abs).resolve()
                relative_output_path = os.path.relpath(abs_output_path, abs_storage_path)
                # Simple check: if it starts with '..' it's likely outside storage/app
                if relative_output_path.startswith(".."):
                     log_warning(f"Output path {abs_output_path} appears outside storage/app. Storing absolute path instead.")
                     relative_output_path = str(abs_output_path) # Store absolute path as string
                else:
                     relative_output_path = relative_output_path.replace('\\', '/') # Ensure forward slashes for consistency
            except ValueError as e:
                log_warning(f"Could not determine relative path for {output_path_abs} (Possibly different drives?). Storing absolute path. Error: {e}")
                relative_output_path = str(Path(output_path_abs).resolve()) # Store absolute path

        # Limit error message length
        if error_msg:
            error_msg = error_msg[:65530] # Max length for TEXT type, leaving some buffer

        sql = """
            UPDATE conversion_jobs
            SET status = %s, output_filepath = %s, error_message = %s, updated_at = NOW()
            WHERE id = %s AND status != 'completed' AND status != 'failed'
        """
        vals = (status, relative_output_path, error_msg, job_id)
        cursor.execute(sql, vals)
        conn.commit()
        log_info(f"DB Update: Job {job_id} status set to '{status}'. Rows affected: {cursor.rowcount}")

    except ValueError as ve:
        log_error(f"DB Configuration Error: {ve}")
    except mysql.connector.Error as err:
        log_error(f"DB Update Error for Job {job_id}: {err}")
    except Exception as e:
        log_error(f"General DB Update Error for Job {job_id}: {e}")
        log_error(traceback.format_exc())
    finally:
        if cursor:
            cursor.close()
        if conn and conn.is_connected():
            conn.close()
            log_debug("DB connection closed.")


# --- Helper Functions (Excel & HTML Processing) ---

SECTION_HEADERS = [ # Case-insensitive matching will be used
    "APPLICATION", "STANDARDS", "CHARACTERISTICS", "CONSTRUCTION",
    "DIMENSIONS", "ELECTRICAL CHARACTERISTICS", "CONDUCTORS",
    "DE-RATING", "CURRENT CARRYING", "VOLTAGE DROP", "FACTORS", "FEATURES", "TECHNICAL DATA"
]

def is_cell_empty(cell):
    """Check if an openpyxl cell is empty or contains only whitespace."""
    return cell is None or cell.value is None or str(cell.value).strip() == ""

def get_cell_value(cell):
    """Safely get cell value as string, handling potential errors."""
    if cell is None or cell.value is None:
        return ""
    try:
        return str(cell.value).strip()
    except Exception:
        return "" # Return empty string on any conversion error

def is_likely_section_header(row_cells):
    """Check if a row looks like a section header (first cell has text, others likely empty)."""
    if not row_cells or is_cell_empty(row_cells[0]):
        return False

    first_cell_text = get_cell_value(row_cells[0]).upper()
    if not any(header in first_cell_text for header in SECTION_HEADERS):
        return False

    # Check if other cells in the row are mostly empty
    empty_count = sum(1 for cell in row_cells[1:] if is_cell_empty(cell))
    # Allow for some non-empty cells if merged, but primarily check first cell and emptiness
    # Threshold: e.g., >80% of remaining cells are empty
    if len(row_cells) > 1 and (empty_count / (len(row_cells) - 1)) < 0.8:
        # Exception: Check if the first cell is merged across many columns
        first_cell_coord = row_cells[0].coordinate
        # This requires checking merged_cells, which we'll do during table extraction
        return True # Tentatively True, refine later

    # It looks like a section header if the first cell matches and others are mostly empty
    return True


def find_merged_cell_value(cell, sheet, merged_ranges):
    """Find the value of a merged cell by looking at the top-left cell of the merge range."""
    if cell.coordinate in sheet.merged_cells:
        for mr_range in merged_ranges:
            if cell.coordinate in mr_range:
                top_left_cell = sheet[mr_range.min_row][mr_range.min_col -1]
                return get_cell_value(top_left_cell)
    return get_cell_value(cell)

def clean_html_content(html_string):
    """Remove VML, MSO comments, and potentially other unwanted tags."""
    if not html_string:
        return ""
    # Remove VML
    html_string = re.sub(r'<(/?)v:[^>]*>', '', html_string, flags=re.IGNORECASE)
     # Remove XML namespaces
    html_string = re.sub(r'<(/?)o:[^>]*>', '', html_string, flags=re.IGNORECASE)
    html_string = re.sub(r'<(/?)w:[^>]*>', '', html_string, flags=re.IGNORECASE)
    # Remove conditional comments
    html_string = re.sub(r'<!--\[if\s+[^\]]+\]>.*?<!\[endif\]-->', '', html_string, flags=re.DOTALL | re.IGNORECASE)
    # Remove empty tags like <p> </p> aggressively? Maybe later.
    # Remove style attributes? Better handled by BS4 later if needed.
    return html_string

def create_bootstrap_table(soup, table_data, sheet, merged_ranges):
    """Creates a Bootstrap 5 styled table from extracted cell data."""
    if not table_data:
        return None

    table = soup.new_tag('table', attrs={'class': 'table table-bordered table-sm'}) # Added table-sm
    tbody = soup.new_tag('tbody')
    table.append(tbody)

    for row_idx, row_cells_data in enumerate(table_data):
        tr = soup.new_tag('tr')
        is_header_row = False # Simple header detection (e.g., all cells bold?) - needs improvement

        # Basic check: if it's the first row and has content, treat as header?
        if row_idx == 0 and any(cell_data['value'] for cell_data in row_cells_data):
            is_header_row = True # Simplistic header detection

        cell_tag = 'th' if is_header_row else 'td'

        for cell_data in row_cells_data:
            if cell_data.get('is_merged_part', False):
                 continue # Skip cells that are part of a merged range but not the top-left

            cell = soup.new_tag(cell_tag)
            # Add colspan and rowspan if present
            if cell_data['colspan'] > 1:
                cell['colspan'] = str(cell_data['colspan'])
            if cell_data['rowspan'] > 1:
                cell['rowspan'] = str(cell_data['rowspan'])

            # Add content (handle potential HTML in cells carefully)
            # For now, treat as text. A more advanced version could parse cell content.
            cell.string = cell_data['value']
            tr.append(cell)

        # Only append row if it contains cells
        if tr.find_all(True, recursive=False):
             tbody.append(tr)

    # Add a thead if we detected a header row
    if is_header_row and tbody.find('tr'):
         thead = soup.new_tag('thead')
         # Move the first row from tbody to thead
         first_row = tbody.find('tr').extract()
         # Change td to th if they weren't already
         for td_cell in first_row.find_all('td'):
              td_cell.name = 'th'
         thead.append(first_row)
         table.insert(0, thead) # Insert thead before tbody

    # Basic cleanup: remove completely empty rows from the final table
    for row in table.find_all('tr'):
         if all(c.get_text(strip=True) == '' for c in row.find_all(['td', 'th'])):
              row.extract()

    if not table.find_all('tr'): # Return None if table becomes empty after cleanup
         return None

    # Wrap table in responsive div
    responsive_div = soup.new_tag('div', attrs={'class': 'table-responsive mb-3'}) # Added margin
    responsive_div.append(table)
    return responsive_div


def extract_images_from_sheet(sheet, output_image_dir, job_id):
    """Extracts images from an openpyxl sheet and saves them."""
    image_paths = {} # Dictionary to store {image_id: relative_path}
    img_counter = 0
    output_image_dir.mkdir(parents=True, exist_ok=True)

    try:
        for image in sheet._images:
            img_counter += 1
            # Generate a unique filename
            img_filename = f"image_{job_id}_{img_counter}{Path(image.ref).suffix or '.png'}" # Use suffix if available
            img_rel_path = f"{OUTPUTS_BASE_DIR_RELATIVE}/{output_image_dir.name}/{img_filename}" # Path relative to storage/app
            img_abs_path = output_image_dir / img_filename

            try:
                # Correct way to access image data might differ slightly based on openpyxl version
                # Try common ways:
                img_data = None
                if hasattr(image, 'data'): # Newer versions might store bytes directly
                     img_data = image.data() # Method call
                elif hasattr(image, '_data'): # Older versions might use _data
                     img_data = image._data()

                if img_data:
                     # Use Pillow to open and save, ensuring format consistency
                     pil_img = PilImage.open(BytesIO(img_data))
                     # Convert to RGB if necessary (e.g., for saving as JPEG)
                     if pil_img.mode == 'RGBA' and img_abs_path.suffix.lower() in ['.jpg', '.jpeg']:
                          pil_img = pil_img.convert('RGB')
                     pil_img.save(img_abs_path)
                     log_info(f"Saved image to: {img_abs_path}")
                     # Store relative path for HTML generation
                     image_paths[image.anchor._from.row] = img_rel_path.replace('\\', '/') # Store by approx row, relative path
                else:
                     log_warning(f"Could not extract data for image {img_counter}")

            except Exception as img_ex:
                log_error(f"Failed to save image {img_counter} to {img_abs_path}: {img_ex}")

    except Exception as e:
        log_error(f"Error during image extraction: {e}")
        log_error(traceback.format_exc())

    log_info(f"Extracted {len(image_paths)} images.")
    return image_paths


# --- Core Conversion Logic ---

def process_excel_to_html(workbook, job_id, output_dir_abs):
    """Processes the Excel workbook and generates standardized HTML."""
    soup = BeautifulSoup(f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel Conversion</title>
    <!-- Bootstrap CSS -->
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/css/bootstrap.min.css" rel="stylesheet" integrity="sha384-rbsA2VBKQhggwzxH7pPCaAqO46MgnOM80zW1RWuH61DGLwZJEdK2Kadq2F9CUG65" crossorigin="anonymous">
    <style>
        body {{ font-family: sans-serif; padding-top: 1rem; }}
        .excel-container {{ max-width: 1200px; margin: auto; }}
        .section-header {{
            font-size: 1.5rem;
            font-weight: bold;
            margin-top: 2rem;
            margin-bottom: 1rem;
            padding-bottom: 0.5rem;
            border-bottom: 2px solid #dee2e6;
            text-align: center; /* Center headers */
        }}
        .table-responsive {{ margin-bottom: 1.5rem; }}
        .table th {{ background-color: #f8f9fa; }}
        img.excel-image {{ max-width: 100%; height: auto; display: block; margin: 1rem auto; }} /* Center images */
        /* Add more custom styles if needed */
    </style>
</head>
<body>
    <div class="container excel-container">
        <!-- Content will be inserted here -->
    </div>
    <!-- Bootstrap JS Bundle (optional) -->
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.2.3/dist/js/bootstrap.bundle.min.js" integrity="sha384-kenU1KFdBIe4zVF0s0G1M5b4hcpxyD9F7jL+jjXkk+Q2h455rYXK/7HAuoJl+0I4" crossorigin="anonymous"></script>
</body>
</html>""", 'html.parser')

    body_container = soup.find('div', class_='excel-container')
    ws = workbook.active # Process only the active sheet for now
    title_tag = soup.find('title')
    if ws.title:
         title_tag.string = ws.title # Set HTML title from sheet name

    # --- Image Extraction ---
    output_image_dir_abs = output_dir_abs / f"job_{job_id}_images"
    image_paths_by_row = extract_images_from_sheet(ws, output_image_dir_abs, job_id)
    log_debug(f"Image paths by row: {image_paths_by_row}")

    # --- Structure Analysis and Content Generation ---
    current_section_title = None
    current_table_data = []
    merged_ranges = ws.merged_cells.ranges # Get merged cells info once

    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        row_cells = list(row) # Convert generator to list

        # Skip completely empty rows (check actual cell values)
        if all(is_cell_empty(cell) for cell in row_cells):
            continue

        # Check for images associated with this row
        if row_idx in image_paths_by_row:
             img_rel_path = image_paths_by_row[row_idx]
             # Make path relative to the *output HTML file location* if needed,
             # but since we store relative to storage/app, and Laravel handles download,
             # the path relative to storage/app root should work.
             img_tag = soup.new_tag('img', attrs={
                  'src': f"/storage/{img_rel_path}", # Path accessible via storage link
                  'class': 'excel-image img-fluid rounded', # Add Bootstrap classes
                  'alt': f'Image from row {row_idx}'
             })
             # Insert image *before* processing the row content if it's not a header
             if not is_likely_section_header(row_cells):
                  # If we have a current table, finish it before the image
                  if current_table_data:
                       table_html = create_bootstrap_table(soup, current_table_data, ws, merged_ranges)
                       if table_html: body_container.append(table_html)
                       current_table_data = []
                  body_container.append(img_tag)
                  log_debug(f"Added image tag for row {row_idx}")


        # Check for section headers
        if is_likely_section_header(row_cells):
            # Finish previous table/section if any
            if current_table_data:
                table_html = create_bootstrap_table(soup, current_table_data, ws, merged_ranges)
                if table_html: body_container.append(table_html)
                current_table_data = []

            current_section_title = find_merged_cell_value(row_cells[0], ws, merged_ranges) # Use merged value
            header_tag = soup.new_tag('h2', attrs={'class': 'section-header'})
            header_tag.string = current_section_title
            body_container.append(header_tag)
            log_info(f"Detected Section Header: {current_section_title}")

             # If an image was for this header row, add it after the header
            if row_idx in image_paths_by_row and img_tag:
                 body_container.append(img_tag)
                 log_debug(f"Added image tag after header for row {row_idx}")


        else: # It's likely a data row (part of a table)
            row_data = []
            col_idx_visual = 0 # Track visual column index considering colspan
            for cell in row_cells:
                 # Check if this cell is part of a merged range
                 is_merged_part = False
                 cr = None # Cell Range
                 if cell.coordinate in ws.merged_cells:
                      for merged_range in merged_ranges:
                           if cell.coordinate in merged_range:
                                cr = merged_range
                                # Only process the top-left cell of a merged range
                                if cell.coordinate != merged_range.coord.split(':')[0]:
                                     is_merged_part = True
                                break

                 colspan = cr.size['columns'] if cr else 1
                 rowspan = cr.size['rows'] if cr else 1
                 cell_value = find_merged_cell_value(cell, ws, merged_ranges) # Get value considering merges

                 row_data.append({
                      'value': cell_value,
                      'coordinate': cell.coordinate,
                      'is_merged_part': is_merged_part,
                      'colspan': colspan,
                      'rowspan': rowspan
                 })
                 col_idx_visual += colspan

            # Add row data to the current table buffer
            # Only add if the row contains *some* data after resolving merges
            if any(cell_info['value'] for cell_info in row_data if not cell_info['is_merged_part']):
                current_table_data.append(row_data)
            else:
                log_debug(f"Skipping row {row_idx} as it seems empty after merge resolution.")


    # Process the last table if buffer is not empty
    if current_table_data:
        table_html = create_bootstrap_table(soup, current_table_data, ws, merged_ranges)
        if table_html: body_container.append(table_html)

    # Final HTML cleanup (using BS4) - Optional, can be enhanced
    # Example: remove empty paragraphs or divs if needed
    for empty_p in body_container.find_all('p'):
         if not empty_p.get_text(strip=True):
              empty_p.decompose()

    # Return the generated HTML as a string
    final_html = str(soup)
    # Optionally run final regex cleanup if needed
    # final_html = clean_html_content(final_html)
    return final_html


def process_excel_to_sql(workbook, job_id):
    """Processes the Excel workbook and generates SQL INSERT statements."""
    sql_statements = []
    ws = workbook.active # Process only the active sheet
    log_info(f"Processing sheet '{ws.title}' for SQL generation.")

    # --- Data Extraction Logic for SQL ---
    # This needs to be heavily customized based on expected Excel structure
    # Example: Assume we want data from the FIRST table found after a "DIMENSIONS" header
    target_table_data = []
    found_header = False
    in_table = False
    merged_ranges = ws.merged_cells.ranges

    for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
        row_cells = list(row)
        if all(is_cell_empty(cell) for cell in row_cells):
            if in_table: # Empty row might signify end of table
                 break
            continue

        first_cell_val = find_merged_cell_value(row_cells[0], ws, merged_ranges).upper()

        if not found_header:
            if "DIMENSIONS" in first_cell_val: # Found the target header
                found_header = True
                log_info("Found 'DIMENSIONS' section header.")
            continue

        # If we found the header, the next non-empty rows are the table
        if found_header:
            # Simple heuristic: Treat first data row as header, rest as data
            row_values = [find_merged_cell_value(cell, ws, merged_ranges) for cell in row_cells if cell.coordinate not in ws.merged_cells or cell.coordinate == ws.merged_cells[cell.coordinate].split(':')[0]] # Get non-merged values
            # Remove empty trailing values
            while row_values and row_values[-1] == "":
                 row_values.pop()
            if row_values: # Only add if row has data
                 target_table_data.append(row_values)
                 in_table = True


    # --- SQL Statement Generation ---
    if len(target_table_data) > 1: # Need at least header + 1 data row
        # Use first row as headers (sanitize column names)
        headers = [re.sub(r'[^a-zA-Z0-9_]', '', h.replace(' ', '_')).lower() for h in target_table_data[0]]
        table_name = f"excel_import_job_{job_id}" # Example table name

        # Basic CREATE TABLE statement (adjust types as needed)
        create_table_sql = f"CREATE TABLE IF NOT EXISTS `{table_name}` (\n"
        create_table_sql += "  `id` INT AUTO_INCREMENT PRIMARY KEY,\n"
        for header in headers:
            # Basic type guessing (improve if needed)
            col_type = "VARCHAR(255)"
            create_table_sql += f"  `{header}` {col_type} NULL,\n"
        create_table_sql += "  `imported_at` TIMESTAMP DEFAULT CURRENT_TIMESTAMP\n"
        create_table_sql += ");"
        sql_statements.append(create_table_sql)
        sql_statements.append("\n-- Data Inserts --\n")


        # Generate INSERT statements
        insert_sql_base = f"INSERT INTO `{table_name}` (`{'`, `'.join(headers)}`) VALUES "
        values_list = []

        for data_row in target_table_data[1:]: # Skip header row
            row_vals_str = []
            for i, value in enumerate(data_row):
                 if i < len(headers): # Ensure we don't have more data than headers
                      # Basic escaping (use proper SQL escaping library if complex data)
                      if value is None:
                           row_vals_str.append("NULL")
                      else:
                           escaped_value = str(value).replace("'", "''") # Simple quote escaping
                           row_vals_str.append(f"'{escaped_value}'")
            if row_vals_str:
                 values_list.append(f"({', '.join(row_vals_str)})")


        # Combine inserts (e.g., 100 rows per statement)
        chunk_size = 100
        for i in range(0, len(values_list), chunk_size):
            chunk = values_list[i:i + chunk_size]
            if chunk:
                 sql_statements.append(insert_sql_base + ",\n".join(chunk) + ";\n")

    else:
        log_warning("Could not find 'DIMENSIONS' table or table has insufficient data for SQL generation.")
        sql_statements.append("-- No suitable table found or table was empty for SQL generation.\n")

    return "".join(sql_statements)

# --- Main Execution ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Process Excel File to HTML or SQL.')
    parser.add_argument('input_path', help='Path to the input Excel file')
    parser.add_argument('output_path', help='Path to save the output file')
    parser.add_argument('job_id', help='Job ID for status updates', type=int)
    parser.add_argument('format', help='Output format (html/sql)', default='html', choices=['html', 'sql'])
    args = parser.parse_args()

    log_info(f"----- Python Script Start -----")
    log_info(f"Job ID: {args.job_id}, Format: {args.format}")
    log_info(f"Input Path: {args.input_path}")
    log_info(f"Output Path: {args.output_path}")
    log_info(f"Working Directory: {os.getcwd()}")
    log_info(f"Loading .env from: {DOTENV_PATH}")

    output_dir_abs = Path(args.output_path).parent
    output_dir_abs.mkdir(parents=True, exist_ok=True) # Ensure output directory exists

    try:
        # --- Input File Check ---
        input_file = Path(args.input_path)
        if not input_file.exists() or not input_file.is_file():
            raise FileNotFoundError(f"Input file not found or is not a file: {args.input_path}")
        if input_file.stat().st_size == 0:
             raise ValueError(f"Input file is empty: {args.input_path}")

        # --- Load Workbook ---
        log_info("Loading workbook...")
        try:
             # Try with read_only=True for potentially faster loading, but may affect image extraction? Test this.
             # data_only=True is crucial to get calculated values instead of formulas.
             workbook = openpyxl.load_workbook(args.input_path, data_only=True, read_only=False) # read_only=False might be safer for images
        except Exception as load_err:
             raise ValueError(f"Failed to load Excel file. It might be corrupted or an unsupported format. Error: {load_err}")

        log_info("Workbook loaded successfully.")

        # --- Process based on format ---
        output_content = None
        if args.format == 'html':
            log_info("Processing for HTML format...")
            output_content = process_excel_to_html(workbook, args.job_id, output_dir_abs)
        elif args.format == 'sql':
            log_info("Processing for SQL format...")
            output_content = process_excel_to_sql(workbook, args.job_id)
        else:
            # Should not happen due to argparse choices, but as a safeguard:
            raise ValueError(f"Unsupported output format: {args.format}")

        # --- Write Output File ---
        if output_content:
             log_info(f"Writing output file to: {args.output_path}")
             try:
                  with open(args.output_path, 'w', encoding='utf-8') as f:
                       f.write(output_content)
                  log_info("Output file written successfully.")
             except Exception as write_err:
                  raise IOError(f"Failed to write output file {args.output_path}. Error: {write_err}")
        else:
             raise ValueError(f"Processing resulted in empty content for format {args.format}.")


        # --- Update Status to Completed ---
        update_job_status_db(args.job_id, 'completed', args.output_path)

        log_info(f"----- Python Script End (Success) -----")
        sys.exit(0) # Success exit code

    # --- Error Handling ---
    except FileNotFoundError as fnf_error:
        error_message = f"Input File Error: {str(fnf_error)}"
        log_error(error_message)
        update_job_status_db(args.job_id, 'failed', error_msg=error_message)
        sys.exit(1)
    except ValueError as val_error: # Specific errors like empty file, load error, bad format
         error_message = f"Processing Error: {str(val_error)}"
         log_error(error_message)
         update_job_status_db(args.job_id, 'failed', error_msg=error_message)
         sys.exit(1)
    except IOError as io_error: # File writing error
         error_message = f"Output Error: {str(io_error)}"
         log_error(error_message)
         update_job_status_db(args.job_id, 'failed', error_msg=error_message)
         sys.exit(1)
    except Exception as e: # Catch-all for other unexpected errors
        error_message = f"Unexpected Python Error: {str(e)}\n{traceback.format_exc()}"
        log_error(error_message)
        update_job_status_db(args.job_id, 'failed', error_msg=error_message[:65530]) # Limit length
        sys.exit(1) # Failure exit code